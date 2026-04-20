
# -*- coding: utf-8 -*-
# ============================================================
# LICEY ROSTER & LINEUP — Google Colab App (No Macros)
# ============================================================
# - Upload an .xlsm/.xlsx
# - Map roster columns
# - Generate lineup + printable outputs
#
# NOTE: This is a starter scaffold. It mirrors Excel features:
#   * Reads named ranges & list validations
#   * Lets you map columns interactively
#   * Builds a basic lineup with simple constraints
# You can expand the "rules" engine to match your full macro logic.
# ============================================================

# 0) INSTALLS (Colab)
try:
    import google.colab  # type: ignore
    IN_COLAB = True
except Exception:
    IN_COLAB = False

if IN_COLAB:
    !pip -q install openpyxl reportlab ipywidgets pandas numpy
    from google.colab import output
    output.enable_custom_widget_manager()

# 1) IMPORTS
import io, json, re, os, datetime as dt
from collections import defaultdict, Counter
import numpy as np
import pandas as pd
from IPython.display import display, HTML
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

# Widgets
try:
    import ipywidgets as widgets
except Exception as e:
    raise RuntimeError("ipywidgets is required. In Colab, run: pip install ipywidgets; then enable widget manager.") from e

# PDF export (simple)
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle

# ============================================================
# 2) HELPERS: Excel Introspection
# ============================================================

def load_excel_bytes_to_wb(file_bytes: bytes):
    # data_only=False so we get formulas where present (if needed).
    return load_workbook(io.BytesIO(file_bytes), data_only=False, keep_vba=True, read_only=False)

def get_defined_map(wb):
    dm = defaultdict(list)  # {name: [(sheet, coord), ...]}
    for dn in wb.defined_names.definedName:
        for title, coord in dn.destinations:
            dm[dn.name].append((title, coord))
    return dm

def get_used_bounds(ws):
    min_row, min_col, max_row, max_col = None, None, 0, 0
    for row in ws.iter_rows():
        for c in row:
            if c.value not in (None, ""):
                r, cidx = c.row, c.column
                if min_row is None or r < min_row: min_row = r
                if min_col is None or cidx < min_col: min_col = cidx
                if r > max_row: max_row = r
                if cidx > max_col: max_col = cidx
    if min_row is None:
        return (0,0,0,0)
    return (min_row, min_col, max_row, max_col)

def ws_to_dataframe(ws):
    # Make a DataFrame assuming first non-empty row is header.
    rmin, cmin, rmax, cmax = get_used_bounds(ws)
    if rmax == 0:
        return pd.DataFrame()
    data = []
    for r in range(rmin, rmax+1):
        row = []
        for c in range(cmin, cmax+1):
            row.append(ws.cell(row=r, column=c).value)
        data.append(row)
    # first row as header
    header = [str(x) if x not in (None,"") else f"Col{idx+1}" for idx,x in enumerate(data[0])]
    df = pd.DataFrame(data[1:], columns=header)
    return df

def extract_validations(wb, defined_map):
    """Return a list of dropdown/list validations with resolved options when possible."""
    out = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.data_validations is None or not ws.data_validations.dataValidation:
            continue
        for dv in ws.data_validations.dataValidation:
            entry = {
                "sheet": sname,
                "type": dv.type,
                "allowBlank": dv.allowBlank,
                "operator": getattr(dv, "operator", None),
                "formula1": dv.formula1,
                "formula2": dv.formula2,
                "targets": str(dv.sqref)
            }
            # Try resolve list values
            vals = []
            if dv.type == "list" and dv.formula1:
                f1 = dv.formula1
                m = re.match(r'^"(.+)"$', f1)
                if m:
                    vals = [x.strip() for x in m.group(1).split(",")]
                else:
                    fr = f1.lstrip("=")
                    if fr in defined_map:
                        for (t_sheet, coord) in defined_map[fr]:
                            rng_min_col, rng_min_row, rng_max_col, rng_max_row = range_boundaries(coord)
                            for r in range(rng_min_row, rng_max_row+1):
                                for c in range(rng_min_col, rng_max_col+1):
                                    v = wb[t_sheet].cell(row=r, column=c).value
                                    if v not in (None, ""):
                                        vals.append(str(v))
                    else:
                        if "!" in fr:
                            t_sheet, coord = fr.split("!", 1)
                            t_sheet = t_sheet.replace("'", "")
                            try:
                                rng_min_col, rng_min_row, rng_max_col, rng_max_row = range_boundaries(coord)
                                for r in range(rng_min_row, rng_max_row+1):
                                    for c in range(rng_min_col, rng_max_col+1):
                                        v = wb[t_sheet].cell(row=r, column=c).value
                                        if v not in (None, ""):
                                            vals.append(str(v))
                            except Exception:
                                pass
            entry["resolved_values_preview"] = list(dict.fromkeys(vals))[:30]
            out.append(entry)
    return out

# ============================================================
# 3) Roster & Rules Model
# ============================================================

REQUIRED_POSITIONS = ["P","C","1B","2B","3B","SS","LF","CF","RF"]  # DH optional
ALL_POSITIONS = REQUIRED_POSITIONS + ["DH","UT","IF","OF"]

def normalize_col(s):
    if s is None: return None
    s2 = str(s).strip().lower()
    s2 = re.sub(r'[^a-z0-9]+', '_', s2)
    return s2

def heuristics_guess_columns(df: pd.DataFrame):
    cols = {normalize_col(c): c for c in df.columns}
    mapping = {
        "player": None,
        "primary_pos": None,
        "secondary_pos": None,
        "bats_throws": None,
        "status": None,
        "role": None,
        "priority": None
    }
    for k in cols:
        if mapping["player"] is None and k in ("player","jugador","name","nombre","player_name"):
            mapping["player"] = cols[k]
        if mapping["primary_pos"] is None and k in ("pos","position","primary_pos","posicion","pos_primary"):
            mapping["primary_pos"] = cols[k]
        if mapping["secondary_pos"] is None and k in ("pos2","secondary_pos","pos_sec","posicion2"):
            mapping["secondary_pos"] = cols[k]
        if mapping["bats_throws"] is None and k in ("bt","b_t","bats_throws","mano","hand"):
            mapping["bats_throws"] = cols[k]
        if mapping["status"] is None and k in ("status","estado","availability","disponible","available"):
            mapping["status"] = cols[k]
        if mapping["role"] is None and k in ("role","rol","paper","papel","tipo"):
            mapping["role"] = cols[k]
        if mapping["priority"] is None and k in ("prio","priority","orden","order","rank","depth"):
            mapping["priority"] = cols[k]
    return mapping

def sanitize_roster(df, map_cfg):
    df2 = df.copy()
    # Required "player" & "primary_pos"
    if not map_cfg.get("player") or not map_cfg.get("primary_pos"):
        raise ValueError("Mapping requires Player and Primary Position columns.")
    # Fill missing columns if not mapped
    for key in ["secondary_pos","bats_throws","status","role","priority"]:
        if key not in map_cfg or map_cfg[key] is None:
            df2[key] = None
        else:
            df2[key] = df2[map_cfg[key]]
    # Rename key columns to standard
    std_cols = {
        "player": map_cfg["player"],
        "primary_pos": map_cfg["primary_pos"],
    }
    for key in ["secondary_pos","bats_throws","status","role","priority"]:
        if map_cfg.get(key):
            std_cols[key] = map_cfg[key]
    df2 = df2.rename(columns=std_cols)
    # Normalize status
    if "status" in df2.columns:
        df2["status"] = df2["status"].astype(str).str.strip().str.lower()
    # Normalize positions
    def norm_pos(x):
        if x is None or (isinstance(x,float) and np.isnan(x)): return None
        s = str(x).strip().upper()
        s = s.replace(" ", "")
        return s
    for col in ["primary_pos","secondary_pos"]:
        if col in df2.columns:
            df2[col] = df2[col].apply(norm_pos)
    # Priority numeric
    if "priority" in df2.columns:
        def to_float(x):
            try:
                return float(x)
            except Exception:
                return np.nan
        df2["priority"] = df2["priority"].apply(to_float)
    return df2

def eligible_for_position(row, pos):
    # Basic eligibility rules; expand as needed.
    if pos == "DH":
        return True
    if row.get("primary_pos") == pos:
        return True
    if row.get("secondary_pos") == pos:
        return True
    # Utility mappings
    if pos in ("LF","CF","RF") and row.get("primary_pos") == "OF":
        return True
    if pos in ("1B","2B","3B","SS") and row.get("primary_pos") == "IF":
        return True
    return False

def build_lineup(roster_df: pd.DataFrame, use_dh=True, exclude_unavailable=True):
    # Filter by status if desired
    r = roster_df.copy()
    if exclude_unavailable and "status" in r.columns:
        mask = ~r["status"].isin(["injured","unavailable","out","na","no"])
        r = r[mask]

    chosen = {}
    used_players = set()

    # Candidate pool sorting
    if "priority" in r.columns and r["priority"].notna().any():
        r_sorted = r.sort_values(by=["priority","player"], ascending=[True, True], na_position="last")
    else:
        r_sorted = r.sort_values(by=["player"])

    positions = REQUIRED_POSITIONS.copy()
    if use_dh:
        positions = ["DH"] + positions  # let DH pick first to avoid conflicts

    for pos in positions:
        # Pick best available eligible player not yet used
        cand = r_sorted[r_sorted.apply(lambda row: eligible_for_position(row, pos), axis=1)]
        cand = cand[~cand["player"].isin(used_players)]
        if len(cand) == 0:
            chosen[pos] = {"player": None, "pos": pos, "note": "No eligible player found"}
            continue
        pick = cand.iloc[0].to_dict()
        chosen[pos] = {"player": pick.get("player"), "pos": pos, "note": ""}
        used_players.add(pick.get("player"))

    # Bench = remaining
    bench = r_sorted[~r_sorted["player"].isin(used_players)].copy()
    lineup_rows = [{"Order": i+1, "POS": p, "Player": chosen[p]["player"]} for i,p in enumerate(positions)]
    lineup_df = pd.DataFrame(lineup_rows)
    return lineup_df, bench

def lineup_to_html(lineup_df: pd.DataFrame, bench_df: pd.DataFrame, title="Lineup Card"):
    css = """
    <style>
    body { font-family: Arial, sans-serif; }
    h2 { margin-bottom: 6px; }
    table { border-collapse: collapse; width: 100%; margin-bottom: 12px; }
    th, td { border: 1px solid #999; padding: 6px 8px; text-align: left; font-size: 12pt; }
    th { background: #eee; }
    .small { font-size: 10pt; color: #444; }
    </style>
    """
    html = [css, f"<h2>{title}</h2>"]
    html.append("<table><thead><tr><th>#</th><th>POS</th><th>Player</th></tr></thead><tbody>")
    for _, row in lineup_df.iterrows():
        html.append(f"<tr><td>{row['Order']}</td><td>{row['POS']}</td><td>{row['Player'] or ''}</td></tr>")
    html.append("</tbody></table>")
    # Bench
    if len(bench_df) > 0:
        html.append("<div class='small'><b>Bench:</b> " + ", ".join([str(x) for x in bench_df['player'].tolist()[:25]]) + "</div>")
    return "\n".join(html)

def save_pdf_lineup(lineup_df: pd.DataFrame, bench_df: pd.DataFrame, pdf_path="lineup_card.pdf", title="Lineup Card"):
    # Simple landscape table PDF
    c = canvas.Canvas(pdf_path, pagesize=landscape(letter))
    width, height = landscape(letter)

    data = [["#", "POS", "Player"]]
    for _, row in lineup_df.iterrows():
        data.append([row["Order"], row["POS"], row["Player"] or ""])

    table = Table(data, colWidths=[50, 80, 400])
    style = TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("FONTSIZE", (0,0), (-1,-1), 12),
        ("ALIGN", (0,0), (0,-1), "CENTER"),
        ("ALIGN", (1,0), (1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ])
    table.setStyle(style)

    c.setFont("Helvetica-Bold", 16)
    c.drawString(40, height-40, title)
    w, h = table.wrapOn(c, width-80, height-120)
    table.drawOn(c, 40, height-80-h)

    # Bench text
    bench_txt = "Bench: " + ", ".join([str(x) for x in bench_df["player"].tolist()[:25]])
    c.setFont("Helvetica", 10)
    c.drawString(40, 40, bench_txt[:200])  # truncate long lists

    c.showPage()
    c.save()
    return pdf_path

# ============================================================
# 4) UI — Upload + Sheet Selection + Column Mapping
# ============================================================

state = {
    "wb": None,
    "sheets": {},
    "defined_map": {},
    "validations": [],
    "roster_df": None,
    "mapping": None,
    "lineup_df": None,
    "bench_df": None
}

file_uploader = widgets.FileUpload(accept=".xlsm,.xlsx", multiple=False, description="Upload Excel")
sheet_dropdown = widgets.Dropdown(options=[], description="Roster Sheet:", disabled=True)

# Column mapping widgets (populated after sheet is chosen)
col_map_widgets = {
    "player": widgets.Dropdown(options=[], description="Player"),
    "primary_pos": widgets.Dropdown(options=[], description="Primary Pos"),
    "secondary_pos": widgets.Dropdown(options=[], description="Secondary Pos"),
    "bats_throws": widgets.Dropdown(options=[], description="Bats/Throws"),
    "status": widgets.Dropdown(options=[], description="Status"),
    "role": widgets.Dropdown(options=[], description="Role"),
    "priority": widgets.Dropdown(options=[], description="Priority")
}

use_dh_chk = widgets.Checkbox(value=True, description="Use DH")
exclude_unavail_chk = widgets.Checkbox(value=True, description="Exclude Unavailable")
build_btn = widgets.Button(description="Generate Lineup", button_style="primary")
export_btn = widgets.Button(description="Export (CSV/HTML/PDF)", button_style="")
logs = widgets.Output()

def on_file_upload(change):
    with logs:
        logs.clear_output()
        if len(file_uploader.value) == 0:
            print("No file uploaded yet.")
            return
        # get bytes
        up = next(iter(file_uploader.value.values()))
        raw = up["content"]
        print(f"Loaded file: {up['metadata']['name']} ({len(raw)} bytes)")
        # parse workbook
        wb = load_excel_bytes_to_wb(raw)
        state["wb"] = wb
        defined_map = get_defined_map(wb)
        state["defined_map"] = defined_map
        # dump sheets to DataFrames
        sheets = {}
        for sname in wb.sheetnames:
            ws = wb[sname]
            df = ws_to_dataframe(ws)
            sheets[sname] = df
        state["sheets"] = sheets
        # validations
        vals = extract_validations(wb, defined_map)
        state["validations"] = vals
        # update sheet dropdown
        sheet_dropdown.options = list(sheets.keys())
        sheet_dropdown.disabled = False
        print(f"Sheets found: {', '.join(wb.sheetnames)}")
        if vals:
            print(f"Detected {len(vals)} data validations. First few list options where resolvable:")
            for v in vals[:5]:
                print(f"  - {v['sheet']} {v['type']} targets={v['targets']} values={v['resolved_values_preview']}")

file_uploader.observe(on_file_upload, names="value")

def on_sheet_select(change):
    with logs:
        logs.clear_output()
        if not change["new"]:
            return
        sname = change["new"]
        df = state["sheets"].get(sname, pd.DataFrame())
        if df.empty:
            print("Selected sheet is empty.")
            return
        state["roster_df"] = df
        print(f"Roster sheet set: {sname}. Columns: {list(df.columns)}")
        # heuristics
        guess = heuristics_guess_columns(df)
        # populate dropdowns
        for key, w in col_map_widgets.items():
            opts = ["<None>"] + list(df.columns)
            w.options = opts
            # preselect guess
            pre = guess.get(key)
            w.value = pre if pre in df.columns else "<None>"
        display(widgets.VBox(list(col_map_widgets.values())))

sheet_dropdown.observe(on_sheet_select, names="value")

def collect_mapping():
    m = {}
    for key, w in col_map_widgets.items():
        m[key] = None if w.value in (None, "<None>") else w.value
    # sanity: player + primary_pos are required
    if not m["player"] or not m["primary_pos"]:
        raise ValueError("You must map at least Player and Primary Position columns.")
    return m

def on_build_clicked(_):
    with logs:
        logs.clear_output()
        if state["roster_df"] is None:
            print("Upload a file and select a roster sheet first.")
            return
        try:
            mapping = collect_mapping()
            df = sanitize_roster(state["roster_df"], mapping)
            state["mapping"] = mapping
            lineup_df, bench_df = build_lineup(
                df,
                use_dh=use_dh_chk.value,
                exclude_unavailable=exclude_unavail_chk.value
            )
            state["lineup_df"] = lineup_df
            state["bench_df"] = bench_df
            print("Lineup built.")
            display(lineup_df)
            print("Bench (first 20):")
            display(bench_df.head(20))
        except Exception as e:
            print("Error generating lineup:", e)

build_btn.on_click(on_build_clicked)

def on_export_clicked(_):
    with logs:
        logs.clear_output()
        if state["lineup_df"] is None or state["bench_df"] is None:
            print("Build a lineup first.")
            return
        lineup_df = state["lineup_df"]
        bench_df = state["bench_df"]

        # CSV
        csv_path = "lineup.csv"
        lineup_df.to_csv(csv_path, index=False, encoding="utf-8-sig")

        # HTML
        html = lineup_to_html(lineup_df, bench_df, title="Game Lineup")
        html_path = "lineup.html"
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html)

        # PDF
        pdf_path = save_pdf_lineup(lineup_df, bench_df, pdf_path="lineup.pdf", title="Game Lineup")

        print("Exported:")
        print(f"- {csv_path}")
        print(f"- {html_path}")
        print(f"- {pdf_path}")
        # Show HTML preview inline
        display(HTML(html))

export_btn.on_click(on_export_clicked)

# Main UI layout
ui = widgets.VBox([
    widgets.HTML("<h3>LICEY Roster & Lineup (Colab)</h3>"),
    file_uploader,
    sheet_dropdown,
    widgets.HBox([use_dh_chk, exclude_unavail_chk]),
    widgets.HBox([build_btn, export_btn]),
    logs
])
display(ui)

