
# -*- coding: utf-8 -*-
# =====================================================================
# LICEY — Minimal Roster & Lineup App (Colab-only, no Excel macros)
# =====================================================================
# Workflow:
#   1) Upload workbook (.xlsm/.xlsx) via Colab's picker.
#   2) App parses "Roster maintenance" sheet and shows roster.
#   3) Edit roster key fields inline (availability, bench, rehab, bats/throws, jersey, BP group).
#   4) Create Home lineup via dropdowns filtered by available players (DH optional).
#   5) Enter Opponent lineup manually.
#   6) Export 3-copy Umpire Cards PDF (both lineups + bench).
# =====================================================================

try:
    import google.colab  # type: ignore
    IN_COLAB = True
except Exception:
    IN_COLAB = False

if IN_COLAB:
    !pip -q install openpyxl reportlab ipywidgets pandas numpy
    from google.colab import output
    output.enable_custom_widget_manager()

import io, re, os, math, datetime as dt
from typing import List, Dict, Tuple
import pandas as pd
import numpy as np
from IPython.display import display, HTML
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
import ipywidgets as widgets

def load_excel_from_bytes(file_bytes: bytes):
    return load_workbook(io.BytesIO(file_bytes), data_only=True, keep_vba=True, read_only=False)

def get_used_bounds(ws):
    min_row, min_col, max_row, max_col = None, None, 0, 0
    for row in ws.iter_rows():
        for c in row:
            if c.value not in (None, ""):
                r, cidx = c.row, c.column
                if min_row is None or r < min_row: min_row = r
                if min_col is None or cidx < min_col: min_col = cidx
                if r > max_row: max_row = r
                if cidx > max_col: max_col = cidx
    if min_row is None:
        return (0,0,0,0)
    return (min_row, min_col, max_row, max_col)

def ws_to_dataframe(ws):
    rmin, cmin, rmax, cmax = get_used_bounds(ws)
    if rmax == 0:
        return pd.DataFrame()
    data = []
    for r in range(rmin, rmax+1):
        row = []
        for c in range(cmin, cmax+1):
            row.append(ws.cell(row=r, column=c).value)
        data.append(row)
    header = [str(x) if x not in (None,"") else f"Col{idx+1}" for idx,x in enumerate(data[0])]
    df = pd.DataFrame(data[1:], columns=header)
    return df

def norm(s):
    if s is None: return ""
    return re.sub(r'[^a-z0-9]+', '_', str(s).strip().lower())

def find_best_match(columns, needles):
    normed = {norm(c): c for c in columns}
    for n in needles:
        if n in normed: 
            return normed[n]
    for c in columns:
        low = str(c).lower()
        for n in needles:
            if n.replace("_"," ") in low or n in norm(c):
                return c
    return ""

def yes_no_to_bool(x):
    if x is None: return False
    s = str(x).strip().lower()
    if s in ("1","true","yes","y","si","sí","available","avail","ok"): return True
    if s in ("0","false","no","n","not","na","unavailable","out"): return False
    return False

def clean_bt(x):
    if x is None: return ""
    s = str(x).strip().upper()
    valid = {"RHS","RHP","LHS","LHP","R","L","S"}
    if s in valid: return s
    s2 = s.replace(" ", "")
    if s2 in valid: return s2
    if "RHP" in s: return "RHP"
    if "LHP" in s: return "LHP"
    if "RHS" in s: return "RHS"
    if "LHS" in s: return "LHS"
    if s.startswith("R"): return "R"
    if s.startswith("L"): return "L"
    if s.startswith("S"): return "S"
    return ""

state = {
    "workbook_name": None,
    "roster_df": pd.DataFrame(),
    "home_lineup": [],
    "opp_lineup": [],
}

logs = widgets.Output()

upload_btn = widgets.Button(
    description="Upload Workbook (.xlsm/.xlsx)",
    button_style="warning",
    tooltip="Uses Colab file picker"
)

roster_box = widgets.VBox([])
save_edits_btn = widgets.Button(description="Save Edits", button_style="success")

use_dh_chk = widgets.Checkbox(value=True, description="Use DH")
build_home_box = widgets.VBox([])

opp_team_txt = widgets.Text(value="", description="Opponent", placeholder="Opponent name")
opp_inputs_box = widgets.VBox([])

export_pdf_btn = widgets.Button(description="Generate Umpire Cards (PDF x3)", button_style="primary")

def parse_workbook_from_bytes(name: str, raw: bytes):
    with logs:
        logs.clear_output()
        print(f"Loaded file: {name} ({len(raw)} bytes)")
    wb = load_excel_from_bytes(raw)
    sheet_name = None
    for s in wb.sheetnames:
        if str(s).strip().lower() == "roster maintenance":
            sheet_name = s
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    df = ws_to_dataframe(ws)

    cols = list(df.columns)
    name_col      = find_best_match(cols, ["name","player","jugador","nombre","player_name"])
    jersey_col    = find_best_match(cols, ["jersey","jersey_number","jersey_#","jersey_no","jersey_num","jersey_"])
    bt_col        = find_best_match(cols, ["handedness","bats","pitchers_handedness_hitters_bats","pitchers_handedness_hitters_bats_r_l_s_","hitters_bats","hand_bat","bats_throws"])
    avail_col     = find_best_match(cols, ["is_the_player_available_today","available","availability","disponible"])
    rehab_col     = find_best_match(cols, ["is_the_player_on_rehab_assignment","rehab"])
    bench_col     = find_best_match(cols, ["is_the_player_currently_on_the_bench","bench"])
    bp_group_col  = find_best_match(cols, ["bp_group","bp","group"])

    missing = []
    for (v,label) in [(name_col,"Name"), (avail_col,"Available")]:
        if not v: missing.append(label)
    if missing:
        with logs:
            print("ERROR: Could not find required columns:", ", ".join(missing))
            print("Headers found:", cols[:50])
        return

    roster = pd.DataFrame({
        "Name": df[name_col],
        "Jersey": df[jersey_col] if jersey_col else "",
        "BT": df[bt_col] if bt_col else "",
        "Available": df[avail_col].apply(yes_no_to_bool) if avail_col else False,
        "Rehab": df[rehab_col].apply(yes_no_to_bool) if rehab_col else False,
        "Bench": df[bench_col].apply(yes_no_to_bool) if bench_col else False,
        "BPGroup": df[bp_group_col] if bp_group_col else ""
    })
    roster["Name"] = roster["Name"].astype(str).str.strip()
    roster["Jersey"] = roster["Jersey"].astype(str).str.replace(".0","", regex=False).str.strip()
    roster["BT"] = roster["BT"].apply(clean_bt)
    roster = roster[roster["Name"].astype(str).str.strip() != ""]
    roster = roster.reset_index(drop=True)

    state["workbook_name"] = name
    state["roster_df"] = roster

    build_roster_editor(roster)
    build_home_lineup_ui()
    build_opp_lineup_ui()

    with logs:
        print(f"Parsed sheet: {sheet_name}")
        print(f"Players loaded: {len(roster)}")

def build_roster_editor(roster: pd.DataFrame):
    rows = []
    header = widgets.HBox([
        widgets.HTML("<b>Name</b>"),
        widgets.HTML("<b>#</b>"),
        widgets.HTML("<b>BT</b>"),
        widgets.HTML("<b>Avail</b>"),
        widgets.HTML("<b>Rehab</b>"),
        widgets.HTML("<b>Bench</b>"),
        widgets.HTML("<b>BPGrp</b>"),
    ])

    editors = []
    for i, r in roster.iterrows():
        name_lbl = widgets.Label(value=str(r["Name"]))
        jersey_in = widgets.Text(value=str(r["Jersey"] or ""), layout=widgets.Layout(width="70px"))
        bt_in = widgets.Dropdown(
            options=["", "R","L","S","RHP","LHP","RHS","LHS"],
            value=str(r["BT"] or ""),
            layout=widgets.Layout(width="100px")
        )
        avail_in = widgets.Checkbox(value=bool(r["Available"]))
        rehab_in = widgets.Checkbox(value=bool(r["Rehab"]))
        bench_in = widgets.Checkbox(value=bool(r["Bench"]))
        bp_in = widgets.Text(value=str(r["BPGroup"] or ""), layout=widgets.Layout(width="70px"))
        row = widgets.HBox([name_lbl, jersey_in, bt_in, avail_in, rehab_in, bench_in, bp_in])
        rows.append(row)
        editors.append((i, name_lbl, jersey_in, bt_in, avail_in, rehab_in, bench_in, bp_in))

    def on_save_edits(_):
        df = state["roster_df"].copy()
        for (i, name_lbl, jersey_in, bt_in, avail_in, rehab_in, bench_in, bp_in) in editors:
            df.at[i, "Jersey"] = jersey_in.value.strip()
            df.at[i, "BT"] = clean_bt(bt_in.value)
            df.at[i, "Available"] = bool(avail_in.value)
            df.at[i, "Rehab"] = bool(rehab_in.value)
            df.at[i, "Bench"] = bool(bench_in.value)
            df.at[i, "BPGroup"] = bp_in.value.strip()
        state["roster_df"] = df
        build_home_lineup_ui()
        with logs:
            print("Roster edits saved. Home lineup options refreshed.")

    save_edits_btn.on_click(on_save_edits)
    roster_box.children = [widgets.HTML("<h4>Roster (Editable)</h4>"), header] + rows + [save_edits_btn]

def build_home_lineup_ui():
    roster = state["roster_df"]
    if roster.empty:
        build_home_box.children = [widgets.Label("Upload a workbook first.")]
        return

    avail_df = roster[roster["Available"] == True].copy()

    def label_of(row):
        j = str(row["Jersey"]).strip()
        return f"{row['Name']} ({j})" if j else f"{row['Name']}"
    avail_df["label"] = avail_df.apply(label_of, axis=1)
    player_options = [""] + avail_df["label"].tolist()

    positions = ["P","C","1B","2B","3B","SS","LF","CF","RF"]
    if use_dh_chk.value:
        positions = ["DH"] + positions

    lineup_rows = []
    dropdowns = []
    for i, pos in enumerate(positions, start=1):
        dd = widgets.Dropdown(options=player_options, description=f"{i}. {pos}", layout=widgets.Layout(width="400px"))
        dropdowns.append((pos, dd))
        lineup_rows.append(dd)

    warn_lbl = widgets.HTML("")
    def validate_unique(_=None):
        picked = [dd.value for _, dd in dropdowns if dd.value]
        dups = [p for p in set(picked) if picked.count(p) > 1]
        warn_lbl.value = "<span style='color:red'>Duplicate player(s): " + ", ".join(dups) + "</span>" if dups else ""

    for _, dd in dropdowns:
        dd.observe(validate_unique, names="value")
    validate_unique()

    def commit_home_lineup():
        lineup = []
        for order, (pos, dd) in enumerate(dropdowns, start=1):
            lineup.append({"Order": order, "POS": pos, "Player": dd.value})
        state["home_lineup"] = lineup

    def on_any_change(_):
        commit_home_lineup()
    use_dh_chk.observe(lambda change: build_home_lineup_ui(), names="value")
    commit_home_lineup()

    build_home_box.children = [widgets.HTML("<h4>Home Lineup (pick from available players)</h4>"), use_dh_chk] + lineup_rows + [warn_lbl]

def build_opp_lineup_ui():
    positions = ["P","C","1B","2B","3B","SS","LF","CF","RF"]
    if use_dh_chk.value:
        positions = ["DH"] + positions

    inputs = []
    rows = []
    for i, pos in enumerate(positions, start=1):
        name_in = widgets.Text(value="", placeholder="Player Name", layout=widgets.Layout(width="350px"))
        row = widgets.HBox([widgets.Label(f"{i}. {pos}", layout=widgets.Layout(width="60px")), name_in])
        rows.append(row)
        inputs.append((pos, name_in))

    def snapshot():
        opp = []
        for order, (pos, name_in) in enumerate(inputs, start=1):
            opp.append({"Order": order, "POS": pos, "Player": name_in.value.strip()})
        state["opp_lineup"] = opp

    for _, name_in in inputs:
        name_in.observe(lambda change: snapshot(), names="value")
    snapshot()

    opp_inputs_box.children = [widgets.HTML("<h4>Opponent Lineup (manual)</h4>"), opp_team_txt] + rows

def parse_label_to_name(label: str) -> str:
    if not label: return ""
    m = re.match(r"^(.*)\s+\(\s*#?.*\s*\)$", label)
    return m.group(1).strip() if m else label

def export_umpire_cards_pdf(path="umpire_cards.pdf"):
    roster = state["roster_df"]
    home = state["home_lineup"]
    opp = state["opp_lineup"]
    opp_team = opp_team_txt.value.strip() or "Opponent"

    def rows_from(lineup, parse_labels=True):
        rows = [["#", "POS", "Player"]]
        for item in lineup:
            player = parse_label_to_name(item["Player"]) if parse_labels else (item["Player"] or "")
            rows.append([item["Order"], item["POS"], player])
        return rows

    home_rows = rows_from(home, parse_labels=True)
    opp_rows  = rows_from(opp, parse_labels=False)

    used_names = set(parse_label_to_name(i["Player"]) for i in home if i["Player"])
    bench_names = []
    for _, r in roster[roster["Available"] == True].iterrows():
        nm = str(r["Name"]).strip()
        if nm and nm not in used_names:
            j = str(r["Jersey"]).strip()
            bench_names.append(f"{nm} ({j})" if j else nm)
    bench_text = ", ".join(bench_names)

    c = canvas.Canvas(path, pagesize=landscape(letter))
    width, height = landscape(letter)

    def draw_page(copy_idx):
        c.setFont("Helvetica-Bold", 16)
        c.drawString(40, height-40, f"Umpire Card — Copy {copy_idx}/3")
        c.setFont("Helvetica", 12)
        today = dt.date.today().strftime("%Y-%m-%d")
        c.drawString(40, height-60, f"Date: {today}")
        c.drawString(180, height-60, f"Opponent: {opp_team}")

        table_home = Table(home_rows, colWidths=[40, 60, 380])
        table_home.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.5, colors.black),
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("FONTSIZE", (0,0), (-1,-1), 12),
            ("ALIGN", (0,0), (0,-1), "CENTER"),
            ("ALIGN", (1,0), (1,-1), "CENTER"),
        ]))
        w, h = table_home.wrapOn(c, width-80, height-140)
        table_home.drawOn(c, 40, height-100-h)

        table_opp = Table(opp_rows, colWidths=[40, 60, 380])
        table_opp.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.5, colors.black),
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("FONTSIZE", (0,0), (-1,-1), 12),
            ("ALIGN", (0,0), (0,-1), "CENTER"),
            ("ALIGN", (1,0), (1,-1), "CENTER"),
        ]))
        w2, h2 = table_opp.wrapOn(c, width-80, height-140)
        table_opp.drawOn(c, width/2 + 20, height-100-h2)

        c.setFont("Helvetica", 11)
        c.drawString(40, 60, "Bench (Available, not in starting lineup):")
        bench_lines = []
        line = ""
        for token in bench_text.split(", "):
            if len(line) + len(token) + 2 > 110:
                bench_lines.append(line)
                line = token
            else:
                line = (line + ", " + token) if line else token
        if line:
            bench_lines.append(line)
        y = 40
        for bl in bench_lines:
            c.drawString(40, y, bl)
            y -= 14

    for i in range(1, 4):
        draw_page(i)
        c.showPage()

    c.save()
    return path

def on_upload_clicked(_):
    with logs:
        logs.clear_output()
        print("Choose your .xlsm/.xlsx in the file picker...")
    try:
        from google.colab import files as gfiles
    except Exception:
        with logs:
            print("Colab upload not available outside Google Colab.")
        return
    result = gfiles.upload()
    if not result:
        with logs:
            print("No file chosen.")
        return
    name = next(iter(result.keys()))
    raw = result[name]
    parse_workbook_from_bytes(name, raw)

upload_btn.on_click(on_upload_clicked)

def on_export_pdf_clicked(_):
    with logs:
        logs.clear_output()
        if state["roster_df"].empty:
            print("Upload and edit roster first.")
            return
        path = export_umpire_cards_pdf("umpire_cards.pdf")
        print("Umpire cards generated:", path)

export_pdf_btn.on_click(on_export_pdf_clicked)

app = widgets.VBox([
    widgets.HTML("<h3>LICEY — Minimal Roster & Lineup</h3>"),
    upload_btn,
    widgets.HTML("<hr>"),
    widgets.HTML("<h4>Step 1 & 2 — Roster</h4>"),
    roster_box,
    widgets.HTML("<hr>"),
    widgets.HTML("<h4>Step 3 — Home Lineup</h4>"),
    build_home_box,
    widgets.HTML("<hr>"),
    widgets.HTML("<h4>Step 4 — Opponent Lineup</h4>"),
    opp_inputs_box,
    widgets.HTML("<hr>"),
    export_pdf_btn,
    logs
])

display(app)
